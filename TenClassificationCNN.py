#分类
import tensorflow as tf
import os
import time
from openpyxl import Workbook #操作excel
from openpyxl import load_workbook
from LoadImage import train_image

#训练集位置
datapath = "E:\\Flow classification\\TransformImage\\train"
#测试集位置
testpath = "E:\\Flow classification\\TransformImage\\test"
#标签存放位置
resultpath = "E:\\Flow classification\\TransformImage\\Result"
#实验次数
exp_count = 44
#excel存放位置
result_excel = "E:\\Flow classification\\TransformImage\\ExpResult"
#模型名字
modelname = "model(10)"
#每个批次大小
batch_size = 15
#计算一共有多少个批次
n_batch = len(os.listdir(datapath))//batch_size
#分类和标签的对应关系
labeldict = {}

#Tensorboard参数概要
def variable_summaries(var):
    with tf.name_scope("summaries"): #解决命名冲突问题
        mean = tf.reduce_mean(var)
        tf.summary.scalar('mean',mean)#平均值
        with tf.name_scope("stddev"):
            stddev = tf.sqrt(tf.reduce_mean(tf.square(var-mean)))#计算平方根
        tf.summary.scalar('stddev',stddev)#标准差
        tf.summary.scalar('max',tf.reduce_max(var))#最大值
        tf.summary.scalar('min',tf.reduce_min(var))#最小值
        tf.summary.histogram('histogram',var)#直方图

#初始化权值
def weight_variable(shape):
    initial = tf.truncated_normal(shape,stddev=0.1)#生成一个截断的正太分布
    return tf.Variable(initial)

#初始化偏置
def bias_variable(shape):
    initial = tf.constant(0.1,shape=shape)#创建常量
    return tf.Variable(initial)

#卷积层
def conv2d_same(x,W):
    #x input tensor of shape '[batch, in_height. in_width. in_channels]' 输入的图片张量
    #W filter / kernel tensor of shape [filter_heiht. filter_width, in_channels, out_channels] 卷积核
    #'strides[0] = strides[3] = 1' strides[1]代表x方向的步长，strides[2]代表y方向的步长
    #padding:A 'string' from:'"SAME", "VALID"' SAME表示考虑边界
    return tf.nn.conv2d(x,W,strides=[1,1,1,1],padding='SAME')

def conv2d_valid(x,W):
    return tf.nn.conv2d(x, W, strides=[1, 1, 1, 1], padding="VALID")

#池化层
def max_pool_2x2(x):
    #ksize[1,x,y,1] 参数依次是：输入(feature map),池化窗口大小，步长，边界
    return tf.nn.max_pool(x,ksize=[1,2,2,1],strides=[1,2,2,1],padding='SAME')


with tf.name_scope('input'):
    x = tf.placeholder(tf.float32,[None,32,32,1],name='x')
    y = tf.placeholder(tf.float32,[None,10],name='y')
    keep_prob = tf.placeholder(tf.float32, name='keep_prob')

with tf.name_scope('learning_rate'):
    LR = tf.Variable(1e-4,dtype=tf.float32)
    tf.summary.scalar("lr", LR)

#改变x的格式转为4D的向量[batch,in_height.in_width, in_channels]
with tf.name_scope('image_reshape'):
    x_image = tf.reshape(x,[-1,32,32,1])
    tf.summary.image("input",x_image)

with tf.name_scope('Conv_layer1'):
    # 初始化第一个卷积层的权值和偏置
    W_conv1 = weight_variable([3,3,1,4])#3*3的采样窗口，4个卷积核从1个平面抽取特征
    b_conv1 = bias_variable([4])#每一个卷积核一个偏置值
    #把x_image和权值进行卷积，在加上偏置值，然后应用relu激活函数
    with tf.name_scope('w_plus_b1'):
        res_conv1 = conv2d_valid(x_image,W_conv1)+b_conv1
        tf.summary.histogram('res_conv1',res_conv1)
    h_conv1 = tf.nn.relu(res_conv1,name='conv1_relu')
    h_pool1 = max_pool_2x2(h_conv1)#运行max_pooling
with tf.name_scope('Conv_layer2'):
    #初始化第二个卷积层的权值和偏置
    W_conv2 = weight_variable([3,3,4,8])#5*5的采样窗口，8个卷积核从4个平面抽取特
    b_conv2 = bias_variable([8])
    #把和权值进行卷积，在加上偏置值，然后应用relu激函数h_pool1
    with tf.name_scope('w_plus_b2'):
        res_conv2 = conv2d_same(h_pool1,W_conv2)+b_conv2
        tf.summary.histogram("res_conv2",res_conv2)
    h_conv2 = tf.nn.relu(res_conv2)
    h_pool2 = max_pool_2x2(h_conv2)#运行max_pooling

    #32*32的图片第一次卷积后还是28*28，第一次池化后边为15*15
    #第二次卷积后为15*15，第二次池化后变为了8*8
    #经过上面的操作后变为8张8*8的平面

#初始化第一个全连接层的权值
with tf.name_scope('FC1'):
    W_fc1 = weight_variable([8*8*8,2048])#上一层有8*8*8个神经院，全连接层有2048个神经元
    b_fc1 = bias_variable([2048])#2048个偏置值

    #把池化层2的输出扁平化为1维
    h_pool2_flat = tf.reshape(h_pool2,[-1,8*8*8])
    #求第一个全连接层的输出
    h_fc1 = tf.nn.relu(tf.matmul(h_pool2_flat,W_fc1) + b_fc1)
    # keep_prob 用来表示神经元的输出概率
    h_fc1_drop = tf.nn.dropout(h_fc1, keep_prob)

with tf.name_scope('FC2'):

    #初始化第二个全连接层
    #进行二十分类
    W_fc2 = weight_variable([2048,10])
    b_fc2 = bias_variable([10])
with tf.name_scope('Pr'):
    #计算输出
    preduction = tf.nn.softmax(tf.matmul(h_fc1_drop,W_fc2)+b_fc2)

#保存模型
saver = tf.train.Saver()
tf.add_to_collection('pred_network',preduction )
#交叉熵代价函数
with tf.name_scope('loss'):
    cross_entorpy = tf.reduce_mean(tf.nn.softmax_cross_entropy_with_logits(labels=y,logits=preduction))
    tf.summary.scalar("loss",cross_entorpy)
#使用AdamOptinizer进行优化
with tf.name_scope('train'):
    train_step = tf.train.AdamOptimizer(LR).minimize(cross_entorpy)
#结果存放在一个bool列表中
with tf.name_scope('accuracy'):
    with tf.name_scope("correct_prediction"):
        correct_prediction = tf.equal(tf.argmax(preduction,1),tf.argmax(y,1))
    #求准确率
    with tf.name_scope('accuracy'):
        accuracy = tf.reduce_mean(tf.cast(correct_prediction,tf.float32))
        tf.summary.scalar('accuracy',accuracy)

#合并记录项
merged = tf.summary.merge_all()

with tf.Session() as sess:
    #初始化参数
    sess.run(tf.global_variables_initializer())
    train_writer = tf.summary.FileWriter('logs/train/',sess.graph)
    start_time = time.time()
    if os.path.exists(result_excel):
        print("存在存放结果的文件夹")
    else:
        os.makedirs(result_excel)
        print("创建文件成功")
    start_make_img_time = time.time()
    #获取训练集和标签
    trainimages, trainlabels, labeldict = train_image(datapath)
    end_make_img_time = time.time()
    print("处理图片用时{}".format(end_make_img_time - start_make_img_time))
    for epoch in range(1,101):
        end_make_img_time=time.time()
        for batch in range(n_batch):
            trainimage,trainlabel =trainimages[batch*batch_size:batch_size*(batch+1)], trainlabels[batch*batch_size:batch_size*(batch+1)]
            summary, _ = sess.run([merged, train_step], feed_dict={x: trainimage, y: trainlabel, keep_prob: 0.7})
        end_train_time = time.time()
        train_writer.add_summary(summary,epoch)
        print("训练用时"+str(end_train_time - end_make_img_time))
        class_count = []
        class_acc = []
        row = []
        row.append(epoch)
        if epoch == 1:
            row_title = ["批次"]+list(labeldict.keys())+['综合准确率']
            if os.path.exists(resultpath):
                with open(resultpath+"\\"+"result"+str(exp_count)+".txt",'w') as f:
                    for line in labeldict.items():
                        ph = line[1].index(max(line[1]))
                        line = line[0]+' '+str(ph)+"\n"
                        f.writelines(line)
            else:
                os.makedirs(resultpath)
                with open(resultpath+"\\"+"result"+str(exp_count)+".txt",'w') as f:
                    for line in labeldict.items():
                        ph = line[1].index(max(line[1]))
                        line = line[0] + ' ' + str(ph) + "\n"
                        f.writelines(line)
            #创建excel保存每次训练各个分类的准确率变化
            wb = Workbook()
            sheet = wb.active
            sheet.append(row_title)
            wb.save(result_excel+"\\result"+str(exp_count)+".xlsx")
        for classname in labeldict.keys():
            #定义测试时相关数据保存的地方
            test_writer = tf.summary.FileWriter('logs/test/'+classname+'/')
            testfaceimage,testlabel = train_image(testpath,classname=classname)
            class_count.append(len(testlabel))
            end_make_test_img_time = time.time()
            summary,one_class_acc = sess.run([merged,accuracy],feed_dict={x:testfaceimage,y:testlabel,keep_prob:1.0})
            end_test_time = time.time()
            # 添加到同一个记录中
            test_writer.add_summary(summary,epoch)
            test_writer.flush()
            print("Tter " + str(epoch) + "类别："+classname + "的测试准确率为" + str(one_class_acc) +  "测试用时" + str(end_test_time - end_make_test_img_time))
            row.append(one_class_acc)
            class_acc.append(one_class_acc)
        testfaceimage, testlabel,_ = train_image(testpath)
        test_writer = tf.summary.FileWriter('logs/test/total/')
        summary, totalacc = sess.run([merged, accuracy],feed_dict={x: testfaceimage, y: testlabel, keep_prob: 1.0})
        #添加到同一个记录中
        test_writer.add_summary(summary, epoch)
        test_writer.flush()
        row.append(totalacc)
        print("Tter " + str(epoch) + "总的测试准确率为" + str(totalacc) )
        wb = load_workbook(result_excel+"\\result"+str(exp_count)+".xlsx")
        sheet = wb.active
        sheet.append(row)
        wb.save(result_excel+"\\result"+str(exp_count)+".xlsx")
        #保存模型
        modelsavepath = "E:\\Flow classification\\TransformImage\\{}\\".format(modelname)
        if os.path.exists(modelsavepath):
            saver.save(sess,"E:\\Flow classification\\TransformImage\\{}\\model.ckpt".format(modelname),epoch)
        else:
            os.makedirs("E:\\Flow classification\\TransformImage\\{}\\".format(modelname))
            saver.save(sess, "E:\\Flow classification\\TransformImage\\{}\\model.ckpt".format(modelname), epoch)
        train_writer.close()
        test_writer.close()
    end_time = time.time()
    print("总用时"+str(end_time-start_time))